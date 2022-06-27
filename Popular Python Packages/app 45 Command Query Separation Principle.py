# Principle of Command or Query
# Our methods/functions should either be a command or query
# Specifically a command that perform an action that changes the state of the system
# Or query that return answer to the caller without changing state or sideeffect


import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]

# command
wb.create_sheet()


# Query
sheet.cell("a1")


# violation
for row in range(1, 10):
    cell = sheet.cell(row, 1)
    print(cell.value)
sheet.append([1, 2, 3])
# Created an unexpected issue, makes the code harder to understand when diagnosing a problem
# Would be good to raise an exception such as

numbers = [1, 2, 3, 4]
numbers[4]

# Trying to access a position that does not exist should raise an exception


wb.save("transactions2.xlsx")
