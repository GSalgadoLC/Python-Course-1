import openpyxl

#wb = openpyxl.Workbook()
wb = openpyxl .load_workbook("transactions.xlsx")
print(wb.sheetnames)


sheet = wb["Sheet1"]

#wb.create_sheet("Sheet2", 0)
# wb.remove_sheet("Sheet2")

cell = sheet["a1"]

print(cell.value)

cell.value = 420
print(cell.row)
print(cell.column)
print(cell.coordinate)

sheet.cell(row=1, column=1)
print(sheet.max_row)
print(sheet.max_column)

for row in range(1, sheet.max_row + 1):
    print("Hi")
    for column in range(1, sheet.max_column+1):
        cell = sheet.cell(row, column)
        print(cell.value)


print()
print("New line")
print()


column = sheet["a"]
print(column)


print()
print("New line")
print()


cells = sheet["a:c"]
print(cells)


print()
print("New line")
print()


atoc3 = sheet["a1:c3"]
print(atoc3)


print()
print("New line")
print()

print(sheet[1:3])


print()
print("New line")
print()


sheet.append([1, 2, 3, 4, 5, 6, 7, 8, 9, 0])
sheet.insert_rows
# delete ./cols ./ rows


print()
print("New line")
print()

wb.save("newFile.xlsx")
